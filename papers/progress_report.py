#!/usr/bin/env python3
"""Generate a progress spreadsheet of the paper collection.

Scans papers/ for PDFs and papers/markdown/ for converted .md files, then
writes an .xlsx with two sheets:

  * "All files" - every source PDF with its classification + conversion status
  * "Paper sets" - a summary of each exam paper (board/spec/series) and how
                   many of its QP / MS / ER / other member files are converted

Python stdlib only (it shells out to a venv that has openpyxl, or you can run
it with any Python that has openpyxl). Usage:

    python3 papers/progress_report.py [out.xlsx]
"""
import argparse
import fnmatch
import re
from collections import defaultdict
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent   # project root
PAPERS = ROOT / "papers"
MARKDOWN = ROOT / "papers" / "markdown"

# ---------------------------------------------------------------- classification

CODE_RE = re.compile(r"(9MA0|8MA0|9FM0|8FM0|9MP0|8MP0)[\s_-]?(\d{1,2})", re.IGNORECASE)
YEAR_RE = re.compile(r"\b(19|20)\d{2}\b")


def classify(name: str) -> dict:
    """Return dict of classification fields for a PDF filename."""
    low = name.lower()
    out = {
        "type": "Other",          # QP, MS, ER, Worked, Insert, Compiled, Skills, Answer, Guide, Spec, Report, Other
        "year": "",
        "code": "",               # exam paper code, e.g. 9MA0/01
        "set": "",
    }

    if "mark scheme" in low:
        out["type"] = "MS"
    elif "examiner" in low or "pef" in low:
        out["type"] = "ER"
    elif "worked" in low or "answers" in low or "_wa" in low:
        out["type"] = "Worked"
    elif "insert" in low:
        out["type"] = "Insert"
    elif "compiled" in low:
        out["type"] = "Compiled"
    elif "skills" in low:
        out["type"] = "Skills"
    elif "specimen" in low or "spec" in low and "specimen" in low:
        out["type"] = "Spec"
    elif "getting started" in low or "guide" in low:
        out["type"] = "Guide"
    elif "subject report" in low or "report" in low:
        out["type"] = "Report"
    else:
        out["type"] = "QP"
        if "-ms" in low or "_ms" in low or ".ms" in low:
            out["type"] = "MS"
        elif "-er" in low or "_er" in low:
            out["type"] = "ER"

    m = CODE_RE.search(name)
    if m:
        out["code"] = f"{m.group(1)}-{m.group(2)}"

    years = list(YEAR_RE.finditer(name))
    if years:
        out["year"] = years[-1].group(0)

    # series label for matching to markdown filenames
    ser = name
    ser = ser.replace(".pdf", "").replace(".PDF", "")
    out["set"] = f"{out['code']}|{out['year']}" if out["code"] and out["year"] else name
    return out


def markdown_exists(board: str, sub: str, set_label: str, kind: str) -> bool:
    """The conversion files are named Edexcel_A-Level_<year>_<paper>.md /
    AEA_AEA_<year>_P1.md etc. Heuristic-by-heuristic matching is messy, so we
    just check whether any .md under markdown/<board>/<sub>/ mentions the same
    code+year/paper. We keep a precomputed index instead."""
    return False


def build_md_index() -> dict:
    """Map (board, sub, code, year, paper) -> md path for every converted file."""
    idx = defaultdict(list)
    for md in MARKDOWN.rglob("*.md"):
        rel = md.relative_to(MARKDOWN).parts  # e.g. ('AEA','AEA_AEA_2020_P1.md')
        if not rel:
            continue
        board = rel[0]
        sub = rel[1] if len(rel) > 2 else ""
        fname = md.stem
        # Dashboard filename conventions we produce:
        #   * AEA:        AEA_AEA_<year>_P1
        #   * Edexcel:    Edexcel_A-Level_<year>_<P#>[-AS][.md]
        info = None
        parts = fname.split("_")
        # AEA_AEA_2020_P1            -> AEA / AEA / 2020 / P1
        # Edexcel_A-Level_2020_S1-AS -> Edexcel / A-Level / 2020 / S1-AS
        if len(parts) >= 4 and re.fullmatch(r"\d{4}", parts[-2]):
            info = {"year": parts[-2], "paper": parts[-1],
                    "path": str(md.relative_to(MARKDOWN))}
        if info:
            idx[board].append(info)
    return idx


def board_type(rel: tuple) -> str:
    """Given a file's rel path under papers/, return (board, subject, series)."""
    if len(rel) >= 2 and rel[0] == "AEA":
        return "AEA", "Mathematics", ""
    if len(rel) >= 2 and rel[0] == "Edexcel":
        return "Edexcel", rel[1], ""
    if len(rel) >= 1 and rel[0] == "MAT":
        return "MAT", "Oxford MAT", ""
    if len(rel) >= 1 and rel[0] == "TMUA":
        return "TMUA", "TMUA", ""
    if len(rel) >= 1 and rel[0] == "markdown":
        return "MARKDOWN", "", ""
    return rel[0] if rel else "", "", ""


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("out", nargs="?", default=str(PAPERS / "papers_progress.xlsx"))
    args = ap.parse_args()

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    # ---------------------------------------------------------------- helpers
    def lookup_md(board: str, subj: str, c: dict, md_index: dict) -> str:
        """Return '✓' if a related conversion exists, else ''."""
        if c["type"] in ("Compiled", "Skills", "Guide", "Insert", "Worked", "Answer", "Report", "Spec"):
            # Support docs aren't individually converted; flag ~ if any related
            # conversion file covers the same code+year.
            hits = [h for h in md_index.get(board, []) if h["year"] == c["year"]]
            return "~" if hits else ""
        # QP/MS/ER: converted if any md covers the same year (board-level)
        hits = [h for h in md_index.get(board, []) if h["year"] == c["year"]]
        return "✓" if hits else ""

    # ---- scan sources ----
    rows = []          # detail rows
    set_rows = defaultdict(lambda: {"n": 0, "converted": 0})
    md_index = build_md_index()

    for pdf in sorted(PAPERS.rglob("*.pdf")):
        rel = pdf.relative_to(PAPERS).parts
        if rel[0] in ("markdown",):
            continue
        board, subj, _ = board_type(rel)
        if not board:
            continue
        c = classify(pdf.name)
        st = lookup_md(board, subj, c, md_index)
        rows.append({
            "path": str(pdf.relative_to(PAPERS)),
            "board": board, "subject": subj,
            "code": c["code"], "year": c["year"], "type": c["type"],
            "converted": st,
            "md": next((h["path"] for h in md_index.get(board, [])
                        if h["year"] == c["year"]
                        and c["type"] == "QP" and h["paper"].startswith(("P", "S", "M", "CP", "FP", "FS", "FM", "FD"))), "")
        })
        key = (board, subj, c["code"] or pdf.stem, c["year"] or pdf.stem)
        set_rows[key]["n"] += 1
        if st in ("✓", "~"):
            set_rows[key]["converted"] += 1

    # ---------------------------------------------------------------- workbook
    wb = Workbook()

    # Sheet 1: detail
    ws = wb.active
    ws.title = "All files"
    headers = ["Path", "Board", "Subject", "Code", "Year", "Type", "Converted", "Markdown file"]
    ws.append(headers)
    for col, w in enumerate([70, 10, 22, 10, 8, 10, 12, 40], start=1):
        ws.column_dimensions[get_column_letter(col)].width = w
    hfill = PatternFill("solid", fgColor="1F4E79")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = hfill
        cell.alignment = Alignment(vertical="center")
    green = Font(color="1E7B1E", bold=True)
    grey = Font(color="888888")
    for r in rows:
        ws.append([r["path"], r["board"], r["subject"], r["code"], r["year"],
                   r["type"], r["converted"], r["md"]])
        rr = ws.max_row
        cell = ws.cell(row=rr, column=7)
        cell.font = green if r["converted"] == "✓" else grey
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Sheet 2: paper-set summary
    ws2 = wb.create_sheet("Paper sets")
    ws2.append(["Board", "Subject", "Code / Name", "Year", "Files", "Converted / related"])
    ws2.column_dimensions["A"].width = 12
    ws2.column_dimensions["B"].width = 22
    ws2.column_dimensions["C"].width = 18
    ws2.column_dimensions["D"].width = 8
    ws2.column_dimensions["E"].width = 8
    ws2.column_dimensions["F"].width = 20
    for cell in ws2[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = hfill
    for (board, subj, code, year), v in sorted(set_rows.items()):
        ws2.append([board, subj, code, year, v["n"], v["converted"]])
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = ws2.dimensions

    out = Path(args.out)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"Wrote {out}")
    print(f"  {len(rows)} source PDF rows, {len(set_rows)} paper sets")


if __name__ == "__main__":
    main()